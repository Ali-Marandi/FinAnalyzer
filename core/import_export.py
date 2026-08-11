"""
Import/Export module for FinAnalyzer Enterprise v2.0.0.
Handles CSV and Excel import with custom column mapping, OFX/QIF bank format parsing,
professional PDF report generation, formatted Excel export with formulas, and JSON API export.
"""

import json
import pandas as pd
from datetime import date
from decimal import Decimal
from typing import Dict, List, Any, Optional
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ofxparse import OfxParser
from qifparse.parser import QifParser

class DataImportExport:
    """Handles multi-format data ingestion, financial reporting (PDF/Excel), and JSON export."""

    @staticmethod
    def parse_csv_excel(file_path: str, column_mapping: Dict[str, str]) -> List[Dict[str, Any]]:
        """Import CSV or Excel file and map columns to standard schema."""
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        elif file_path.endswith(('.xls', '.xlsx')):
            df = pd.read_excel(file_path)
        else:
            raise ValueError("Unsupported file format. Use CSV or Excel.")

        # Rename columns according to mapping
        df = df.rename(columns=column_mapping)
        return df.to_dict(orient='records')

    @staticmethod
    def parse_ofx(file_path: str) -> List[Dict[str, Any]]:
        """Parse OFX bank statement format."""
        with open(file_path, 'r', encoding='latin1') as f:
            ofx = OfxParser.parse(f)
        
        transactions = []
        for account in ofx.accounts:
            for tx in account.statement.transactions:
                transactions.append({
                    "date": tx.date.date(),
                    "amount": Decimal(str(tx.amount)),
                    "payee": tx.payee,
                    "memo": tx.memo,
                    "id": tx.id
                })
        return transactions

    @staticmethod
    def parse_qif(file_path: str) -> List[Dict[str, Any]]:
        """Parse QIF bank statement format."""
        with open(file_path, 'r', encoding='latin1') as f:
            qif = QifParser.parse(f)

        transactions = []
        for account in qif.get_accounts():
            for tx in account.transactions:
                transactions.append({
                    "date": tx.date,
                    "amount": Decimal(str(tx.amount)),
                    "payee": tx.payee,
                    "memo": tx.memo
                })
        return transactions

    @staticmethod
    def generate_pdf_report(report_title: str, data: Dict[str, Any], output_path: str) -> str:
        """Generate a publication-grade PDF financial report using ReportLab."""
        doc = SimpleDocTemplate(output_path, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=22,
            textColor=colors.HexColor("#1A365D"),
            spaceAfter=15,
            alignment=1
        )
        
        story.append(Paragraph(report_title, title_style))
        story.append(Spacer(1, 10))

        # Add summary table or key metrics
        table_data = [["Metric / Account", "Amount / Value"]]
        for k, v in data.items():
            table_data.append([str(k).replace('_', ' ').title(), f"{v:,.2f}" if isinstance(v, (int, float, Decimal)) else str(v)])

        t = Table(table_data, colWidths=[300, 200])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2B6CB0")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F7FAFC")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E0")),
        ]))
        story.append(t)

        doc.build(story)
        return output_path

    @staticmethod
    def export_excel_with_formulas(financial_data: List[Dict[str, Any]], output_path: str) -> str:
        """Export formatted financial statement to Excel with automated SUM formulas."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Statement"

        # Styling
        header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ["Account Code", "Account Name", "Debit", "Credit", "Net Balance"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        row_idx = 2
        for item in financial_data:
            ws.append([
                item.get("code"),
                item.get("name"),
                float(item.get("debit", 0)),
                float(item.get("credit", 0)),
                f"=C{row_idx}-D{row_idx}"
            ])
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = border
            row_idx += 1

        # Add total row with SUM formula
        ws.append(["Total", "", f"=SUM(C2:C{row_idx-1})", f"=SUM(D2:D{row_idx-1})", f"=SUM(E2:E{row_idx-1})"])
        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = Font(bold=True)
            cell.border = border

        wb.save(output_path)
        return output_path

    @staticmethod
    def export_json_api(data: Any) -> str:
        """Export data to JSON format for API consumption."""
        return json.dumps(data, default=str, indent=4)
